"""
Excel出力モジュール

このモジュールでは、処理結果をExcel形式で出力するための機能を提供します。
Excel生成の基本機能、カスタムヘッダー設定、データ変換、スタイル適用などの機能が含まれます。
"""

import os
import logging
import tempfile
from pathlib import Path
from typing import List, Dict, Any, Optional, BinaryIO
from datetime import datetime

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

from ..data.models import ProcessResult, ExcelConfig
from ..data.interfaces import ExcelExporterProtocol, ProcessResultProtocol
from ..utils.errors import ExcelExportError, with_error_handling


class ExcelExporter(ExcelExporterProtocol):
    """
    Excel出力クラス
    
    処理結果をExcel形式で出力します。
    Excel生成の基本機能、カスタムヘッダー設定、データ変換、スタイル適用などの機能が含まれます。
    """
    
    def __init__(self, config: ExcelConfig):
        """
        初期化
        
        Args:
            config: Excel出力設定
        """
        self.logger = logging.getLogger(__name__)
        self.config = config
    
    @with_error_handling(ExcelExportError, "Excel出力処理でエラーが発生しました")
    def export(self, results: List[ProcessResultProtocol], output_path: Path) -> Path:
        """
        処理結果をExcel形式でエクスポートします。
        
        Args:
            results: 処理結果のリスト
            output_path: 出力ファイルのパス
            
        Returns:
            エクスポートされたファイルのパス
            
        Raises:
            ExcelExportError: Excel出力処理でエラーが発生した場合
        """
        self.logger.info(f"Excel出力開始: 結果数={len(results)}, 出力先={output_path}")
        
        # 出力ディレクトリが存在しない場合は作成
        output_path.parent.mkdir(parents=True, exist_ok=True)
        
        # 既存ファイルのバックアップ
        if output_path.exists():
            backup_path = self._create_backup(output_path)
            self.logger.info(f"既存ファイルをバックアップしました: {backup_path}")
        
        # 新しいワークブックを作成
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = "スタイルタイトル"
        
        # ヘッダーの設定
        self._set_headers(sheet)
        
        # データの追加
        self._add_data(sheet, results)
        
        # 列幅の自動調整
        self._adjust_column_widths(sheet)
        
        # スタイルの適用
        self._apply_styles(sheet, len(results))
        
        # 保存
        try:
            workbook.save(output_path)
            self.logger.info(f"Excelファイルを保存しました: {output_path}")
            return output_path
        except Exception as e:
            self.logger.error(f"Excelファイルの保存エラー: {e}")
            raise ExcelExportError(f"Excelファイルの保存に失敗しました: {str(e)}", 
                                 output_path=str(output_path)) from e
    
    @with_error_handling(ExcelExportError, "Excelバイナリデータの生成でエラーが発生しました")
    def get_binary_data(self, results: List[ProcessResultProtocol]) -> bytes:
        """
        処理結果のExcelバイナリデータを取得します。
        
        Args:
            results: 処理結果のリスト
            
        Returns:
            Excelバイナリデータ
            
        Raises:
            ExcelExportError: Excel出力処理でエラーが発生した場合
        """
        self.logger.info(f"Excelバイナリデータ生成開始: 結果数={len(results)}")
        
        # 新しいワークブックを作成
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = "スタイルタイトル"
        
        # ヘッダーの設定
        self._set_headers(sheet)
        
        # データの追加
        self._add_data(sheet, results)
        
        # 列幅の自動調整
        self._adjust_column_widths(sheet)
        
        # スタイルの適用
        self._apply_styles(sheet, len(results))
        
        # 一時ファイルに保存してバイナリデータを取得
        try:
            with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as tmp:
                tmp_path = tmp.name
            
            workbook.save(tmp_path)
            
            with open(tmp_path, 'rb') as f:
                binary_data = f.read()
            
            # 一時ファイルを削除
            os.unlink(tmp_path)
            
            self.logger.info(f"Excelバイナリデータを生成しました: {len(binary_data)} バイト")
            return binary_data
            
        except Exception as e:
            self.logger.error(f"Excelバイナリデータの生成エラー: {e}")
            # 一時ファイルが残っていたら削除
            if 'tmp_path' in locals() and os.path.exists(tmp_path):
                os.unlink(tmp_path)
            raise ExcelExportError(f"Excelバイナリデータの生成に失敗しました: {str(e)}") from e
    
    def _create_backup(self, file_path: Path) -> Path:
        """
        ファイルのバックアップを作成します。
        
        Args:
            file_path: バックアップするファイルのパス
            
        Returns:
            バックアップファイルのパス
        """
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        backup_path = file_path.with_name(f"{file_path.stem}_{timestamp}_backup{file_path.suffix}")
        
        # ファイルをコピー
        import shutil
        shutil.copy2(file_path, backup_path)
        
        return backup_path
    
    def _set_headers(self, sheet: openpyxl.worksheet.worksheet.Worksheet) -> None:
        """
        シートにヘッダーを設定します。
        
        Args:
            sheet: 設定対象のワークシート
        """
        headers = self.config.headers
        
        for col_letter, header_text in headers.items():
            cell = sheet[f"{col_letter}1"]
            cell.value = header_text
    
    def _add_data(self, sheet: openpyxl.worksheet.worksheet.Worksheet, results: List[ProcessResultProtocol]) -> None:
        """
        シートにデータを追加します。
        
        Args:
            sheet: 追加対象のワークシート
            results: 処理結果のリスト
        """
        for i, result in enumerate(results, start=2):  # ヘッダー行の次から
            # A列: スタイリスト名
            sheet[f"A{i}"] = result.selected_stylist.name
            
            # B列: クーポン名
            sheet[f"B{i}"] = result.selected_coupon.name
            
            # C列: コメント
            sheet[f"C{i}"] = result.selected_template.comment
            
            # D列: スタイルタイトル
            sheet[f"D{i}"] = result.selected_template.title
            
            # E列: 性別
            sheet[f"E{i}"] = result.attribute_analysis.sex
            
            # F列: 長さ
            sheet[f"F{i}"] = result.attribute_analysis.length
            
            # G列: スタイルメニュー
            sheet[f"G{i}"] = result.selected_template.menu
            
            # H列: ハッシュタグ
            sheet[f"H{i}"] = result.selected_template.hashtag
            
            # I列: 画像ファイル名
            sheet[f"I{i}"] = result.image_name
    
    def _adjust_column_widths(self, sheet: openpyxl.worksheet.worksheet.Worksheet) -> None:
        """
        列幅を自動調整します。
        
        Args:
            sheet: 調整対象のワークシート
        """
        for column in sheet.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            
            # ヘッダーセルの幅に対応するために係数を調整
            adjusted_width = (max_length + 2) * 1.1
            
            # 最大幅と最小幅を設定
            min_width = 10
            max_width = 50
            adjusted_width = max(min_width, min(max_width, adjusted_width))
            
            sheet.column_dimensions[column_letter].width = adjusted_width
    
    def _apply_styles(self, sheet: openpyxl.worksheet.worksheet.Worksheet, row_count: int) -> None:
        """
        シートにスタイルを適用します。
        
        Args:
            sheet: 適用対象のワークシート
            row_count: データ行数
        """
        # ヘッダー行のスタイル
        header_font = Font(bold=True, size=12, color="FFFFFF")
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        
        # セル枠線
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # ヘッダー行にスタイルを適用
        for col in range(1, len(self.config.headers) + 1):
            cell = sheet.cell(row=1, column=col)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border
        
        # データ行のスタイル
        data_alignment = Alignment(vertical="center", wrap_text=True)
        
        # データ行にスタイルを適用
        for row in range(2, row_count + 2):  # ヘッダーの次からデータ行まで
            for col in range(1, len(self.config.headers) + 1):
                cell = sheet.cell(row=row, column=col)
                cell.alignment = data_alignment
                cell.border = thin_border
                
                # 特定のカラムに特別なスタイルを適用
                # 例：ハッシュタグカラム（H列）のスタイル
                if col == 8:  # H列
                    cell.alignment = Alignment(vertical="center", wrap_text=True)
